#! python3
# -*- coding: utf-8 -*-

# 演習プロジェクト14.8.1は、ExcelからCSVに変換するプログラムを作るものであるが、
# サンプルとして与えられているデータexcelSpreadsheets.zipの内容が、
# Excelファイルでなく、spreadsheet-A.csv ～ spreadsheet-Z.csv であったため、
# 逆に、これらからExcelファイルを作成するプログラムを作りました。
#
#    python cxv2xlsx spreadsheet
#
# とすると、spreadsheet.xlsx ができます。

import openpyxl
import sys
import os
import csv
import re

if len(sys.argv) < 2:
    sys.exit('Usage: python csv2xlsx.py prefix')

prefix = sys.argv[1]

wb = openpyxl.Workbook()
wb.remove_sheet(wb.active)

for csv_filename in sorted(os.listdir('.')):
    mo = re.match(r'(.+?)[\-_](.+?)\.csv', csv_filename)
    if not mo or mo.group(1) != prefix:
        continue
    sheet_name = mo.group(2)
    sheet = wb.create_sheet(sheet_name)

    csv_file = open(csv_filename, 'r')
    csv_reader = csv.reader(csv_file)
    for n,row in enumerate(csv_reader):
        for m,col in enumerate(row):
            sheet.cell(column=m+1, row=n+1).value = col
    csv_file.close()

wb.save(prefix + '.xlsx')

